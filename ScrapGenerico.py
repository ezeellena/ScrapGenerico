import datetime
import json
import pickle
import urllib
import requests
import csv
import re
import os
from unidecode import unidecode
from bs4 import BeautifulSoup
from openpyxl import Workbook
from nube import texto2els


def actualizar_tema(grupo_telegram, tema_anterior):

    try:
        grupo_telegram = grupo_telegram.lower()


        # me conecto a la API de Sonda para ver que tema tiene asignado cada grupo
        url = url_api_sonda + "/sb_get_tema/?grupo="+grupo_telegram
        r = requests.get(url).text

        if tema_anterior != ",".join(eval(r)):
            enviar_noticias(["---------------------------------------------------------------\n"
                             "--"
                             "\n *El término de búsqueda ahora es:* "+ r +""
                             "\n------------------------------------------------------------"])

        return r
    except Exception as e:
        return "-1"

j_link_enviado = {}
j_pag_ne = {}
mugre = ["xmlns=http://www.w3.org/1999/>", "<\n", "\n>", "<<p>", "<p>", "</p", "xmlns=http://www.w3.org/1999/>",
         "xmlns=http://www.w3.org/1999/>", "<br />", "CDATA", "</div>>", "<div>", "</div>", "%>", "<iframe>",
         "</iframe>", "100%", "<div", "http://w3.org/", "xmlms", "xhtml", ";>", "<", ">", "'", '"', "\/", "]", "["]
def limpiar(texto, mugre):
    for m in mugre:
        texto = texto.replace(m, "")
    return texto
def link_enviado(l):
    l = limpiar(l, mugre)
    if not l in j_link_enviado.keys():
        print(" Enviando a telegram :", l)
        j_link_enviado[l] = 1
        return False
    else:
        print(" !!!!!!!!!!!!!!! ENCONTRO DUPLICADO !!!!!!!!!!!!!!!!!! ")
        return True
def log(texto):
    l = open("log.csv", "a")
    l.write(texto + "\n")
    l.close()
def save_persist(elem):
    try:
        vpath = "./persist/"

        varchivo = vpath + elem + ".bin"
        with open(varchivo, "bw") as archivo:
            pickle.dump(eval(elem), archivo)
    except Exception as e:

        print("Except de save_persist", e)
def load_persist(elem):
    try:

        vpath = "./persist/"
        varchivo = vpath + elem + ".bin"

        with open(varchivo, "br") as archivo:
            # #print(pickle.load(archivo))
            return pickle.load(archivo)

    except Exception as e:

        print("269 - Except load_persit ", e)
def replaceBase(texto):
    texto = texto.replace("http:", " ").replace("//", " ").replace(".", " ").replace("www", " ").replace(
            "https:", " ").replace("/", " ").replace("\n", " ").replace("-", " ")
    return  texto
def replaceURL(texto):
    texto = texto.replace("http:", "").replace("//", "").replace(".", "").replace("www", "").replace(
            "https:", "").replace("/", "").replace("\n", "").replace("-", "")
    return  texto
def contarElementosLista(lista):
    """
    Recibe una lista, y devuelve un diccionario con todas las repeticiones decada valor
    """
    return {i:lista.count(i) for i in lista}
def configuracionExcels(url):
    global PagNoticiaLink
    PagNoticiaLink = Workbook()
    global hoja
    hoja = PagNoticiaLink.active
    hoja['A1'] = "URL"
    hoja['A2'] = url
    hoja['A3'] = "TEXTO OBTENIDO"
    hoja['B3'] = "LINKS DE LAS NOTICIAS"
    hoja['C3'] = "TAG CON EL QUE ENCONTRO LAS NOTICIAS"
    hoja['D3'] = "TODOS LOS LINKS DE LA NOTICIA"
    hoja['E3'] = "HTML"
    hoja['F3'] = "TITULO NOTICIA"
    hoja['G3'] = "DESCRIPCION NOTICIA"
    hoja['H3'] = "FECHA PUBLICACION NOTICIA"
    hoja['I3'] = "FECHA MODIFICACION NOTICIA"
    hoja['J3'] = "ERRORES"
    hoja['K3'] = "ERRORES"
    hoja['L3'] = "ERRORES"
    hoja['M3'] = "ERRORES"
    hoja['N3'] = "ERRORES"
    hoja['O3'] = "ERRORES"
    hoja['P3'] = "ERRORES"
    hoja['Q3'] = "ERRORES"
    return PagNoticiaLink, hoja

class RSSParser(object):
    def parse(self,confiTagPage, urlytag, tema):
        ListaDeLinks = []
        items = []
        Noticia = []
        RedesSociales = ["facebook", "twitter", "whatsapp"]
        self.url = urlytag
        global urlCortada
        if "news.google" in urlytag:
             urlCortada = "Google"
        else:

            urlCortada= replaceURL(urlytag)
        LINKS = open("./LINKS/" + urlCortada + ".csv", "a", encoding='utf-8')

        configuracionExcels(url)
        try:
            headers = {
                'User-Agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10.12; rv:55.0) Gecko/20100101 Firefox/55.0',
            }
            response = requests.get(url, headers=headers).text
        except Exception as e:
            print("Error 1 - Obtener Response ", e)
        fila = 0
        try:
            LINKS.write('----------LINK-----------' + '\n' + str(url))
            for Noti in confiTagPage["j"]["BuscarNoticia"]:
                fila += 1
                try:
                    Noticias = eval(Noti)
                    if Noticias != []:
                        Noticia.extend(Noticias)
                        for i, Noticiae in enumerate(Noticias):
                            LIIINKS = [a['href'] for a in Noticiae.find_all('a', href=True)]
                            LIIINKS = ', '.join(LIIINKS)
                            hoja.cell(row=i + 4, column=1).value = filtroReplace(unidecode(Noticiae.text))
                            hoja.cell(row=i + 4, column=4).value = LIIINKS
                            hoja.cell(row=i + 4, column=3).value = str(Noti)
                            LINKS.write('----------HTML-----------:'+ '\n' + str(filtroReplace(Noticiae.text)) + '\n')
                except Exception as e:
                    hoja.cell(row=fila + 4, column=10).value = str(e.args)
                    hoja.cell(row=fila + 4, column=11).value = filtroReplace(Noticiae.text)
                    print("Error 2 - Obtener Articulos de noticias ", e)
        except Exception as e:
            print("Error 3 - Obtener Articulos de noticias ", e)
        temp9 = ""
        row = 3

        try:
            for i in Noticia:
                row +=1

                hoja.cell(row=row, column=5).value = str(i)
                texto = filtroReplace(i.get_text())
                if filtro_tema2(texto, tema) and texto != '':

                    ListaDeLinks = eval(confiTagPage["j"]["path"])
                    if ListaDeLinks != "":
                        SetDeLinks = set(ListaDeLinks)
                        url2 = url
                        try:
                            resultado = contarElementosLista(ListaDeLinks)
                            maximo = max(resultado, key=resultado.get)
                            print("El valor mas repetido es el ", maximo, " con ", resultado[maximo], " veces")
                        except Exception as e:
                            print(" 4 - Obtener Resultado maximo de links ", e)
                            hoja.cell(row=fila + 4, column=12).value = str(e)
                            hoja.cell(row=fila + 4, column=13).value = str(resultado)
                            hoja.cell(row=fila + 4, column=14).value = str(maximo)
                        if len(SetDeLinks) == 1 and list(SetDeLinks)[0] != urlytag:
                            temp9 = list(SetDeLinks)
                            temp9 = str(temp9[0])
                            hoja.cell(row=row, column=2).value = temp9
                            LINKS.write('----------LINK-----------'+'\n'+ str(unidecode(temp9))+'\n'+  '----------HTML-----------:' +'\n'+ unidecode(str(texto)) +  '\n')
                        else:
                            if resultado[maximo] >= 2:
                                temp9 = maximo
                                hoja.cell(row=row, column=2).value = temp9
                                LINKS.write('----------LINK-----------' + '\n' + str(unidecode(
                                    temp9)) + '\n' + '----------HTML-----------:' + '\n' + unidecode(str(
                                    texto)) + '\n')
                            else:
                                palabras = texto.split()
                                palabras.append(tema)
                                try:
                                    for l in ListaDeLinks:
                                        linkCortado = replaceBase(l).split()
                                        totalPalabras = len([palabra for palabra in palabras if palabra in linkCortado])
                                        if 2 <= totalPalabras <= 20:
                                            totalRedes = len([redSocial for redSocial in RedesSociales if redSocial in l.lower()])
                                            if not totalRedes >= 1:
                                                temp9 = l
                                                LINKS.write(unidecode(temp9) + '\n')

                                except Exception as e:
                                    hoja.cell(row=fila + 4, column=15).value = str(e.args)
                                    hoja.cell(row=fila + 4, column=16).value = str(i)
                                    print(" ERROR 5 - NO ESCRAPEO NADA ", e)
                                    print(" ********* URL no parseada correctamente: \n", url, "\n")
                                    print(i)
                                    print("**********************************************************")

                                    if not i.text in j_pag_ne.keys():
                                        j_pag_ne[i.text] = 1
                                        log("***** \n No scrapeó esta página: \n" + url + '\n' + str(i) + '\n ********')


                            """
                            print("- url: ", url)
                            print("- temp9:" , temp9)
                            print("--------------------------------")
                        """
                    if temp9[:1] == "/":
                        temp9 = temp9[1:]
                    if temp9[:2] == "./":
                        temp9 = temp9[2:]
                    if temp9[:8] == "noticias":
                        temp9 = temp9[8:]
                    if "http" in temp9:
                        url2 = temp9
                        temp9 = ""

                            # print(" url final:  ", url +  temp9 + temp10 + temp11, "\n temp9: ", temp9, "\n temp10: ",temp10, "\n temp11: ",temp11, "\n temp12: ",temp12)
                    if "news.google" in url2:
                        url2 = "https://news.google.com/"
                    j_i = {"link": url2 + temp9,
                           "desc": texto,
                           "tmpItems": i}
                    if not filtro_repetida(j_i):
                        archivoCSV = []
                        LinkNotcia = j_i["link"]
                        DescripcionNoticia = ""
                        fechaPublicacion = ""
                        response2 = requests.get(LinkNotcia, headers=headers).text
                        Titulo = []
                        for Titu in confiTagPage["j"]["tituloNoticia"]:
                            try:
                                Titulos = eval(Titu)
                                if Titulos != []:
                                    Titulo.append(Titulos)
                            except Exception as e:
                                print(" Obtener Titulo", e,Titu)
                                hoja.cell(row=row, column=17).value = str(e)
                        resultadoTitulo = contarElementosLista(Titulo)
                        maximoTitulo = max(resultadoTitulo, key=resultadoTitulo.get)
                        hoja.cell(row=row, column=6).value = maximoTitulo
                        print("El valor mas repetido es el ", maximoTitulo, " con ", resultadoTitulo[maximoTitulo], " veces")
                        Descripcion = []
                        for Descr in confiTagPage["j"]["descripcionNoticia"]:
                            try:
                                Descripciones = eval(Descr)
                                if Descripciones != []:
                                    Descripcion.append(Descripciones)
                            except Exception as e:
                                print(" Obtener Titulo", e,Descr)
                                hoja.cell(row=row, column=18).value = str(e)
                        resultadoDescripcion = contarElementosLista(Descripcion)
                        maximoDescripcion = max(resultadoDescripcion, key=resultadoDescripcion.get)
                        hoja.cell(row=row, column=7).value = maximoDescripcion
                        print("El valor mas repetido es el ", maximoDescripcion, " con ", resultadoDescripcion[maximoDescripcion], " veces")
                        link_web = url
                        FechaHoraScrapeo = str(datetime.datetime.now())
                        archivoCSV.append(link_web + ',\n' + ',\n' + LinkNotcia + ',\n' + FechaHoraScrapeo)
                        items.append(LinkNotcia)
                        with open("out.csv", "a") as f:
                            wr = csv.writer(f, delimiter="\n")
                            for ele in items:
                                wr.writerow([ele + ","])
                        texto2els(texto, FechaHoraScrapeo, LinkNotcia)
            LINKS.close()
            PagNoticiaLink.save('./Excel/' + urlCortada + '-Noticias.xlsx')
            #hojaExcelDeErrores.save('./Excel/errores' + urlCortada + '-Errores.xlsx')
            return items
        except Exception as e:
            print(" 100 - Obtener links ", e)
            hoja.cell(row=row, column=10).value = str(e)


def filtroReplace(object):
    object.replace("/", "").replace(":", "").replace("%", "").replace("-", "").replace("[", "").replace("]","").replace("<","").replace(">", "").replace("!", "").replace(",", "")
    return " ".join(object.split())
def filtro_repetida(j_i):
    dd = j_i['link'].replace("\n", "")[1:250]

    if "reconquista" in dd:
        print("parar")

    dd = limpiar(dd, mugre)
    # dd = dd.replace("/", "").replace(":", "").replace("%", "").replace("-", "").replace("[", "").replace("]", "").replace("<", "").replace(">", "").replace("!", "").replace("\n","")

    # print(dd)
    r = False
    if dd in db_noticias2.keys():
        if (db_noticias2[dd] == 1):
            r = True
        else:
            r = False
    if not r:
        db_noticias2[dd] = 1
        save_persist('db_noticias2')
    return r
def filtro_tema(j_i, tema):
    c1 = tema.upper() in j_i['desc'].upper()
    for j in tema.split(","):
        c2 = j.upper() in j_i['desc'].upper()
        c1 = c1 or c2
    if c1:
        r = True
    else:
        r = False
    return r
def filtro_tema2(texto, tema):
    c1 = tema.upper() in texto.upper()
    for j in tema.split(","):
        c2 = j.upper() in texto.upper()
        c1 = c1 or c2
    if c1:
        r = True
    else:
        r = False
    return r
a_url_temas = []
def init():
    global vtelegram
    vtelegram = True

    def get_config(vkey):
        c = vkey in j_config.keys()
        if c:
            return j_config[vkey]
        else:
            return '-1'
def enviar_noticias(arr):
    if not vtelegram:
        pass
        # return

    try:
        url_api = token + "/sendMessage"

        # print( "- tema \n", Tema, " \n ",  NombreGrupo )
        men_t = "✔ Noticias referidas al tema %s, enviadas al grupo de télegram %s: " % (Tema, NombreGrupo) + "\n"
        ta = False
        # recorro el arreglo de links y lo imprimos
        men = []
        # print("arr \n",  arr)
        # print( "men_t \n", men_t )
        for a in arr:
            # print("3- ",a)
            # men += "- " + a + "\n\n"

            # armo la linea
            l = "- " + a + "\n\n"
            men.append(l)
            if a != "" and not (link_enviado(a)):
                ta = True
        if ta:
            # Si tiene información, mando el título.
            requests.post('https://api.telegram.org/' + url_api, data={'chat_id': idchat, 'text': men_t})
            for m in men:
                requests.post('https://api.telegram.org/' + url_api,
                              data={'chat_id': idchat, 'text': '\n [' + NombreGrupo + ']\n' + m})
                print(requests.status_codes)
    except Exception as e:
        print(" 279 - enviar ", e)
def configuracion():
    f = open("configParametrosGenerico.json", "r")

    global j_config
    j_config = {}
    j_config = json.loads(f.read())
    global vtelegram
    global grupo_telegram_fijo
    grupo_telegram_fijo = j_config["grupo_telegram_fijo"]
    global url_api_sonda
    url_api_sonda = j_config["url_api_sonda"]

    try:
        vtelegram = j_config["telegram"]
    except:
        vtelegram = True
    global token
    token = j_config["token"]
    global idchat
    idchat = j_config["id_chat"]
    global Tema
    Tema = j_config["Tema"]
    global NombreGrupo
    NombreGrupo = j_config["NombreGrupo"]
    global directorio
    directorio = j_config["directorio"]
    try:
        global db_noticias2
        db_noticias2 = {}
        if os.path.isfile(directorio + 'persist/db_noticias2.bin'):
            db_noticias2 = load_persist("db_noticias2")
        else:
            db_noticias2 = {}
    except:
        db_noticias2 = {}

    try:
        global LinksDePaginasWeb
        LinksDePaginasWeb = {}
        if os.path.isfile(directorio + 'persist/LinksDePaginasWeb.bin'):
            LinksDePaginasWeb = load_persist("LinksDePaginasWeb")
        else:
            LinksDePaginasWeb = {}
    except:
        LinksDePaginasWeb = {}
def write_json(data, filename='configGenerico.json'):
    with open(filename,'w') as f:
        json.dump(data, f, indent=4)

if __name__ == "__main__":
    configuracion()
    j = open("configGenerico.json", "r")
    confiTagPage = {}
    confiTagPage = json.loads(j.read())
    urlNueva = ""
    if urlNueva != "":
        confiTagPage["j"]["link"].append(urlNueva)
        write_json(confiTagPage)
    result = actualizar_tema(grupo_telegram_fijo, Tema)
    if result != "-1":
        Tema = ",".join(eval(result))
    while True:
        try:
            for url in confiTagPage["j"]["link"]:
                r = False
                if url in LinksDePaginasWeb.keys():
                    if (LinksDePaginasWeb[url] == 1):
                        r = True
                    else:
                        r = False
                if not r:
                    LinksDePaginasWeb[url] = 1
                    save_persist('LinksDePaginasWeb')
                if url != "":

                    ################################################################

                    # Actualizo el tema de búsqueda
                    result = actualizar_tema(grupo_telegram_fijo, Tema)
                    if result != "-1":
                        Tema = ",".join(eval(result))

                    # print( "********************************" )
                    print("TEMA:\n\n", Tema)
                    # print( "********************************" )
                    if url == "http://www.paralelo28.com.ar/":
                        eze = 1+1
                    ##########################################################33
                    print(" Procesando la url:  ", url)
                    r = RSSParser().parse(confiTagPage, url, Tema)
                    PagNoticiaLink.save('./Excel/' + urlCortada + '-Noticias.xlsx')
                    if r != []:
                        enviar_noticias(r)
        except Exception as e:
            print("207 - problema en el for general ", e)
    fic.close()
